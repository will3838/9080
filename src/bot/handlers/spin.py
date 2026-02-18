"""Handler for /spin command."""

from __future__ import annotations

import asyncio
import io
import random
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import ContextTypes

from bot.handlers.captcha import generate_captcha
from bot.storage.inventory_db import add_item


async def spin_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return

    user = update.effective_user
    user_id = user.id if user is not None else 0

    spinning_users = context.application.bot_data.setdefault("spinning_users", set())
    if user_id in spinning_users:
        await update.message.reply_text("рулетка уже крутиться. подождите")
        return

    spinning_users.add(user_id)

    try:
        spin_counts = context.application.bot_data.setdefault("spin_counts", {})
        captcha_pending = context.application.bot_data.setdefault("captcha_pending", {})

        if user_id in captcha_pending:
            question = captcha_pending[user_id]["question"]
            await update.message.reply_text(f"сначала реши капчу.\nКапча: {question} Ответь числом.")
            return

        current_count = spin_counts.get(user_id, 0)
        if current_count > 0 and current_count % 50 == 0:
            question, answer = generate_captcha()
            captcha_pending[user_id] = {"question": question, "answer": answer}
            await update.message.reply_text(f"Капча: {question} Ответь числом.")
            return

        items = context.application.bot_data.get("items")
        if not items:
            await update.message.reply_text("Произошла ошибка. Попробуйте позже.")
            return

        weights = [item["chance"] for item in items]
        selected_item = random.choices(items, weights=weights, k=1)[0]

        timestamp = datetime.now(timezone.utc).isoformat()
        chat = update.effective_chat

        username = user.username if user is not None and user.username is not None else ""
        chat_id = chat.id if chat is not None else 0

        item_id = selected_item["id"]
        png_path = Path("item") / f"{item_id}.png"
        jpg_path = Path("item") / f"{item_id}.jpg"

        existing_images = [path for path in (png_path, jpg_path) if path.exists()]
        if len(existing_images) != 1:
            raise RuntimeError(f"Expected exactly one image for item id {item_id}")

        image_path = existing_images[0]
        if image_path.name not in {f"{item_id}.png", f"{item_id}.jpg"}:
            raise RuntimeError(f"Invalid image filename for item id {item_id}")

        try:
            add_item(
                user_id=user_id,
                item_id=item_id,
                username=username,
                chat_id=chat_id,
                timestamp=timestamp,
            )
        except Exception:
            await update.message.reply_text("временная ошибка сохранения, обратитесь к @HATE_death_ME")
            return

        animation_info = context.application.bot_data.get("roulette_animation")
        if animation_info is None:
            raise RuntimeError("Roulette animation is missing")

        animation_buffer = io.BytesIO(animation_info["bytes"])
        animation_buffer.name = animation_info["filename"]
        anim_msg = await update.message.reply_animation(
            animation=animation_buffer,
            duration=animation_info["duration_s"],
            width=animation_info["width"],
            height=animation_info["height"],
        )

        await asyncio.sleep(7)

        try:
            await anim_msg.delete()
        except Exception:
            await update.message.reply_text("не удалось удалить анимацию (нет прав)")

        spin_file = Path("data") / "spin.xlsx"

        if not spin_file.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "spin"
            ws["A1"] = "timestamp"
            ws["B1"] = "user_id"
            ws["C1"] = "username"
            ws["D1"] = "chat_id"
            ws["E1"] = "item_id"
            ws["F1"] = "item_name"
            ws["G1"] = "price"
            ws["H1"] = "chance"
            wb.save(spin_file)

        wb = load_workbook(spin_file)
        ws = wb["spin"] if "spin" in wb.sheetnames else wb.active
        ws.append(
            [
                timestamp,
                user_id,
                username,
                chat_id,
                selected_item["id"],
                selected_item["name"],
                selected_item["price"],
                selected_item["chance"],
            ]
        )
        wb.save(spin_file)

        with image_path.open("rb") as image_file:
            await update.message.reply_photo(
                photo=image_file,
                caption=f"вам выпало {selected_item['name']}",
            )

        spin_counts[user_id] = current_count + 1
    finally:
        spinning_users.discard(user_id)
