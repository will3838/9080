"""Bot entrypoint."""

from __future__ import annotations

import logging
from decimal import Decimal
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from telegram.ext import Application, CommandHandler

from bot.config import load_settings
from bot.errors.error_handler import on_error
from bot.handlers.help import help_command
from bot.handlers.spin import spin_command
from bot.logging_config import setup_logging


logger = logging.getLogger(__name__)


def ensure_item_excel_exists() -> None:
    data_dir = Path("data")
    item_dir = Path("item")
    file_path = data_dir / "item.xlsx"

    data_dir.mkdir(parents=True, exist_ok=True)
    item_dir.mkdir(parents=True, exist_ok=True)

    if file_path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "item"
    ws["A1"] = "id"
    ws["B1"] = "name"
    ws["C1"] = "price"
    ws["D1"] = "chance"
    wb.save(file_path)


def ensure_spin_excel_exists() -> None:
    data_dir = Path("data")
    file_path = data_dir / "spin.xlsx"

    data_dir.mkdir(parents=True, exist_ok=True)

    if file_path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "spin"
    ws["A1"] = "timestamp"
    ws["B1"] = "user_id"
    ws["C1"] = "username"
    ws["D1"] = "chat_id"
    ws["E1"] = "item_id"
    ws["F1"] = "item_name"
    ws["G1"] = "price"
    ws["H1"] = "chance"
    wb.save(file_path)


def load_items_once() -> list[dict[str, object]]:
    file_path = Path("data") / "item.xlsx"

    if not file_path.exists():
        raise RuntimeError("data/item.xlsx does not exist.")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    if "item" not in wb.sheetnames:
        wb.close()
        raise RuntimeError("Worksheet 'item' is missing in data/item.xlsx.")

    ws = wb["item"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    expected_header = ("id", "name", "price", "chance")

    if header != expected_header:
        wb.close()
        raise RuntimeError("Invalid header in data/item.xlsx. Expected: id, name, price, chance.")

    items: list[dict[str, object]] = []

    for row in rows:
        if row is None or len(row) < 4:
            wb.close()
            raise RuntimeError("Invalid data row in data/item.xlsx.")

        item_id, name, price, chance = row[0], row[1], row[2], row[3]

        if not isinstance(item_id, int) or isinstance(item_id, bool):
            wb.close()
            raise RuntimeError("Invalid item id in data/item.xlsx.")

        if not isinstance(name, str) or not name.strip():
            wb.close()
            raise RuntimeError("Invalid item name in data/item.xlsx.")

        if not isinstance(price, int) or isinstance(price, bool):
            wb.close()
            raise RuntimeError("Invalid item price in data/item.xlsx.")

        if not isinstance(chance, (float, Decimal)):
            wb.close()
            raise RuntimeError("Invalid item chance type in data/item.xlsx.")

        chance_value = float(chance)
        if not (0 < chance_value < 1):
            wb.close()
            raise RuntimeError("Invalid item chance value in data/item.xlsx.")

        items.append(
            {
                "id": item_id,
                "name": name,
                "price": price,
                "chance": chance_value,
            }
        )

    wb.close()

    if not items:
        raise RuntimeError("data/item.xlsx must contain at least one data row.")

    return items


def main() -> None:
    """Configure and run the bot via long polling."""
    load_dotenv()
    setup_logging()
    ensure_item_excel_exists()
    ensure_spin_excel_exists()
    items = load_items_once()

    settings = load_settings()

    application = Application.builder().token(settings.telegram_bot_token).build()
    application.bot_data["items"] = items
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("spin", spin_command))
    application.add_error_handler(on_error)

    logger.info("Starting bot with long polling")
    application.run_polling()


if __name__ == "__main__":
    main()
